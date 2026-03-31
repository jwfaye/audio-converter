"""Business logic for converting between Excel and WAV formats.

This module contains the core conversion functions shared by the CLI
(:mod:`audio_converter.cli`) and the GUI (:mod:`audio_converter.gui`).
It has no dependency on ``argparse``, ``print``, or ``sys.exit()``.

The main entry points are:

* :func:`excel_to_wav` -- single-sheet Excel (int16 samples) to WAV.
* :func:`wav_to_excel` -- WAV to single-sheet Excel.
* :func:`wav_to_pipeline_excel` -- WAV through the auditory periphery
  pipeline, producing a multi-sheet Excel with selectable stages
  (audio, peaks, memo_ma/ck, integrated memo_ma/ck).
"""

from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path

import numpy as np
from openpyxl import Workbook, load_workbook

from audio_converter.audio_io import load_audio, save_audio

INT16_MAX = 32768.0
EXCEL_MAX_COLS = 16384


class ConversionError(Exception):
    """Raised when a conversion cannot complete.

    Typical causes include missing input files, unreadable formats,
    write permission errors, or a missing ``soundperception`` install.
    """


@dataclass
class ConversionResult:
    """Result of a successful conversion.

    Attributes:
        output_path: Absolute or relative path to the written file.
        num_samples: Total number of audio samples in the output.
        sample_rate: Sample rate of the audio in Hz.
    """

    output_path: Path
    num_samples: int
    sample_rate: int


def excel_to_wav(
    input_path: Path,
    output_path: Path,
    sample_rate: int,
) -> ConversionResult:
    """Convert an Excel file (first row of samples) to a WAV file.

    Args:
        input_path: Path to the input .xlsx file.
        output_path: Path to the output .wav file.
        sample_rate: Sample rate in Hz.

    Returns:
        ConversionResult with details of the written file.

    Raises:
        ConversionError: If no data is found or conversion fails.
    """
    try:
        wb = load_workbook(input_path, read_only=True, data_only=True)
    except Exception as exc:
        raise ConversionError(f"Impossible d'ouvrir le fichier Excel : {exc}") from exc

    audio_sheets = sorted(
        [s for s in wb.sheetnames if s.startswith("audio")],
        key=lambda s: (int(s.split("_")[1]) if "_" in s else 0),
    )
    if not audio_sheets:
        audio_sheets = [wb.sheetnames[0]]

    samples: list[float] = []
    for sheet_name in audio_sheets:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            for val in row:
                if val is not None:
                    samples.append(float(val))
    wb.close()

    if not samples:
        raise ConversionError(
            "Aucune donnée trouvée dans la première ligne du fichier Excel."
        )

    audio_data = np.array(samples, dtype=np.float32)
    audio_data = audio_data / INT16_MAX

    save_audio(output_path, audio_data, sample_rate)

    return ConversionResult(
        output_path=output_path,
        num_samples=len(samples),
        sample_rate=sample_rate,
    )


def wav_to_excel(
    input_path: Path,
    output_path: Path,
) -> ConversionResult:
    """Convert a WAV file to an Excel file (one row of samples).

    The audio is always resampled to 16 kHz internally.

    Args:
        input_path: Path to the input .wav file.
        output_path: Path to the output .xlsx file.

    Returns:
        ConversionResult with details of the written file.

    Raises:
        ConversionError: If the WAV file cannot be read or conversion fails.
    """
    try:
        audio_data, sr = load_audio(input_path, target_sample_rate=16000)
    except Exception as exc:
        raise ConversionError(f"Impossible de lire le fichier WAV : {exc}") from exc

    wb = Workbook()
    wb.remove(wb.active)

    int16_data = (audio_data * INT16_MAX).astype(np.int16)
    _write_audio_sheets(wb, int16_data)

    try:
        wb.save(output_path)
    except Exception as exc:
        raise ConversionError(f"Impossible d'écrire le fichier Excel : {exc}") from exc

    return ConversionResult(
        output_path=output_path,
        num_samples=len(audio_data),
        sample_rate=sr,
    )


def wav_to_pipeline_excel(
    input_path: Path,
    output_path: Path,
    export_audio: bool = True,
    export_peaks: bool = False,
    export_periphery: bool = True,
    export_integration: bool = False,
    tau: int = 200,
    decimation_factor: int = 100,
    progress_callback: Callable[[str], None] | None = None,
) -> ConversionResult:
    """Convert a WAV file to a multi-sheet Excel with auditory pipeline stages.

    The audio is resampled to 16 kHz and scaled to int16 before being fed
    to :class:`~soundperception.audition.core.periphery.AuditoryPeriphery`.
    Channel order is reversed so that high frequencies appear in the
    first rows and low frequencies at the bottom of each sheet.

    Each enabled stage produces one or more Excel sheets:

    * ``audio`` -- raw int16 signal (1 row per chunk of 16 384 samples).
    * ``peaks`` -- peak-extracted signal after Stage 4
      (65 channels x n_samples).  When both *export_audio* and
      *export_peaks* are ``True``, the audio is appended as the last row
      of the ``peaks`` sheet instead of a separate sheet.
    * ``memo_ma`` / ``memo_ck`` -- auditory nerve amplitude and temporal
      memory (65 channels x n_samples).
    * ``memo_ma_int`` / ``memo_ck_int`` -- temporally integrated and
      decimated versions of memo_ma / memo_ck
      (65 channels x n_decimated).

    Args:
        input_path: Path to the input ``.wav`` file.
        output_path: Path to the output ``.xlsx`` file.
        export_audio: Include raw audio data.
        export_peaks: Include the peaks sheet (Stage 4 output).
        export_periphery: Include ``memo_ma`` and ``memo_ck`` sheets.
        export_integration: Include temporally integrated sheets.
        tau: Temporal integration time constant in samples.
        decimation_factor: Decimation factor for temporal integration.
        progress_callback: Called with a short status string at each
            processing step.  Useful for updating a GUI progress label.

    Returns:
        A :class:`ConversionResult` with output path, sample count, and
        sample rate.

    Raises:
        ConversionError: If the WAV cannot be read, the Excel cannot be
            written, no stage is selected, or ``soundperception`` is not
            installed.
    """
    def _report(msg: str) -> None:
        if progress_callback:
            progress_callback(msg)

    try:
        from soundperception.audition.config import (
            AuditoryPeripheryConfig,
            TemporalIntegrationConfig,
        )
        from soundperception.audition.core.integration import TemporalIntegrator
        from soundperception.audition.core.periphery import AuditoryPeriphery
    except ImportError as exc:
        raise ConversionError(
            "Le module soundperception n'est pas installé. "
            "Installez-le avec : pip install -e ../sound-perception"
        ) from exc

    _report("Chargement audio...")
    try:
        audio_data, sr = load_audio(input_path, target_sample_rate=16000)
    except Exception as exc:
        raise ConversionError(f"Impossible de lire le fichier WAV : {exc}") from exc

    wb = Workbook()
    wb.remove(wb.active)

    if export_audio and not export_peaks:
        _report("Export audio brut...")
        int16_data = (audio_data * INT16_MAX).astype(np.int16)
        _write_audio_sheets(wb, int16_data)

    needs_pipeline = export_peaks or export_periphery or export_integration
    if needs_pipeline:
        _report("Calcul périphérie auditive...")
        config = AuditoryPeripheryConfig.default()
        config.cochlear.sample_rate = sr
        periphery = AuditoryPeriphery(config)
        signal_int16 = (audio_data * INT16_MAX).astype(np.int16)
        result = periphery.process(signal_int16, return_intermediate=True)

        for key in ("peaks", "memo_ma", "memo_ck"):
            if key in result:
                result[key] = np.flipud(result[key])

        if export_peaks:
            _report("Export peaks...")
            peaks_sheets = _write_2d_array(wb, "peaks", result["peaks"])
            if export_audio:
                _report("Export audio brut (sous peaks)...")
                int16_data = (audio_data * INT16_MAX).astype(np.int16)
                _append_audio_row(wb, peaks_sheets, int16_data)

        if export_periphery:
            _report("Export memo MA...")
            _write_2d_array(wb, "memo_ma", result["memo_ma"])

            _report("Export memo CK...")
            _write_2d_array(wb, "memo_ck", result["memo_ck"])

        if export_integration:
            _report("Calcul intégration temporelle...")
            int_config = TemporalIntegrationConfig(
                tau=tau,
                decimation_factor=decimation_factor,
            )
            integrator = TemporalIntegrator(int_config)
            int_result = integrator.process(result["memo_ma"], result["memo_ck"])

            _report("Export memo MA intégré...")
            _write_2d_array(wb, "memo_ma_int", np.flipud(int_result["memo_ma"]))

            _report("Export memo CK intégré...")
            _write_2d_array(wb, "memo_ck_int", np.flipud(int_result["memo_ck"]))

    if not wb.sheetnames:
        raise ConversionError("Aucune étape sélectionnée pour l'export.")

    _report("Écriture fichier Excel...")
    try:
        wb.save(output_path)
    except Exception as exc:
        raise ConversionError(f"Impossible d'écrire le fichier Excel : {exc}") from exc

    return ConversionResult(
        output_path=output_path,
        num_samples=len(audio_data),
        sample_rate=sr,
    )


def _write_audio_sheets(wb: Workbook, int16_data: np.ndarray) -> None:
    """Write int16 audio samples across one or more ``audio`` sheets.

    Excel columns are limited to :data:`EXCEL_MAX_COLS` (16 384), so
    longer signals are split across sheets named ``audio``,
    ``audio_2``, ``audio_3``, etc.

    Args:
        wb: Target workbook (sheets are created in-place).
        int16_data: 1-D array of int16 audio samples.
    """
    chunks = [
        int16_data[i : i + EXCEL_MAX_COLS]
        for i in range(0, len(int16_data), EXCEL_MAX_COLS)
    ]
    for idx, chunk in enumerate(chunks):
        name = "audio" if idx == 0 else f"audio_{idx + 1}"
        ws = wb.create_sheet(name)
        ws.append([int(s) for s in chunk])


def _append_audio_row(
    wb: Workbook, sheet_names: list[str], int16_data: np.ndarray
) -> None:
    """Append int16 audio samples as a row across the given sheets.

    The samples are split into chunks of :data:`EXCEL_MAX_COLS` and
    appended to the corresponding sheets so that audio and 2-D data
    stay aligned.

    Args:
        wb: Workbook containing the target sheets.
        sheet_names: Sheet names (as returned by :func:`_write_2d_array`).
        int16_data: 1-D array of int16 audio samples.
    """
    for idx, name in enumerate(sheet_names):
        col_start = idx * EXCEL_MAX_COLS
        col_end = min(col_start + EXCEL_MAX_COLS, len(int16_data))
        if col_start >= len(int16_data):
            break
        wb[name].append([int(s) for s in int16_data[col_start:col_end]])


def _write_2d_array(
    wb: Workbook, sheet_name: str, array: np.ndarray
) -> list[str]:
    """Write a 2-D NumPy array across one or more sheets.

    When the number of samples (columns) exceeds :data:`EXCEL_MAX_COLS`,
    the array is split column-wise across sheets named ``sheet_name``,
    ``sheet_name_2``, ``sheet_name_3``, etc.

    Args:
        wb: Target workbook (sheets are created in-place).
        sheet_name: Base name for the created sheet(s).
        array: 2-D array of shape ``(n_channels, n_samples)``.

    Returns:
        List of sheet names that were created.
    """
    _n_channels, n_samples = array.shape
    sheet_names: list[str] = []
    for chunk_idx, col_start in enumerate(range(0, n_samples, EXCEL_MAX_COLS)):
        col_end = min(col_start + EXCEL_MAX_COLS, n_samples)
        name = sheet_name if chunk_idx == 0 else f"{sheet_name}_{chunk_idx + 1}"
        ws = wb.create_sheet(name)
        for row in array[:, col_start:col_end]:
            ws.append([float(v) for v in row])
        sheet_names.append(name)
    return sheet_names
