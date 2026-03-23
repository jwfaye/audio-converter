"""Tests for audio-converter: round-trip, multi-sheet, soundperception import."""

import numpy as np
import pytest
from pathlib import Path
from openpyxl import load_workbook

from audio_converter.audio_io import save_audio, load_audio
from audio_converter.converter import (
    ConversionError,
    ConversionResult,
    excel_to_wav,
    wav_to_excel,
    wav_to_pipeline_excel,
    EXCEL_MAX_COLS,
    INT16_MAX,
)


# ── Helpers ───────────────────────────────────────────────────────────────


def _make_wav(tmp_path: Path, samples: np.ndarray, sr: int = 16000) -> Path:
    """Write a float32 WAV file and return its path."""
    wav_path = tmp_path / "test.wav"
    save_audio(wav_path, samples, sr)
    return wav_path


def _sine(duration_s: float, sr: int = 16000, freq: float = 440.0) -> np.ndarray:
    """Generate a sine wave in [-1, 1]."""
    t = np.arange(int(sr * duration_s), dtype=np.float32) / sr
    return (0.9 * np.sin(2 * np.pi * freq * t)).astype(np.float32)


# ── soundperception import ────────────────────────────────────────────────


class TestSoundPerceptionImport:
    def test_import_soundperception(self):
        """soundperception must be importable (deployment smoke test)."""
        import soundperception  # noqa: F401

    def test_import_auditory_periphery(self):
        from soundperception.audition.config import AuditoryPeripheryConfig  # noqa: F401
        from soundperception.audition.core.periphery import AuditoryPeriphery  # noqa: F401

    def test_import_temporal_integration(self):
        from soundperception.audition.config import TemporalIntegrationConfig  # noqa: F401
        from soundperception.audition.core.integration import TemporalIntegrator  # noqa: F401


# ── Round-trip: short audio (<= 1s, single sheet) ────────────────────────


class TestRoundTripShort:
    """Round-trip tests with audio that fits in one Excel sheet."""

    def test_wav_to_excel_to_wav(self, tmp_path):
        """WAV → Excel → WAV produces matching audio."""
        original = _sine(0.5)
        wav_in = _make_wav(tmp_path, original)
        xlsx = tmp_path / "out.xlsx"
        wav_out = tmp_path / "roundtrip.wav"

        r1 = wav_to_excel(wav_in, xlsx)
        assert r1.output_path == xlsx
        assert r1.num_samples == len(original)

        r2 = excel_to_wav(xlsx, wav_out, sample_rate=16000)
        assert r2.num_samples == len(original)

        recovered, sr = load_audio(wav_out)
        assert sr == 16000
        assert len(recovered) == len(original)
        np.testing.assert_allclose(recovered, original, atol=1e-4)

    def test_excel_to_wav_to_excel(self, tmp_path):
        """Excel → WAV → Excel preserves int16 samples."""
        original = _sine(0.5)
        wav_in = _make_wav(tmp_path, original)
        xlsx1 = tmp_path / "step1.xlsx"
        wav_mid = tmp_path / "mid.wav"
        xlsx2 = tmp_path / "step2.xlsx"

        wav_to_excel(wav_in, xlsx1)
        excel_to_wav(xlsx1, wav_mid, sample_rate=16000)
        wav_to_excel(wav_mid, xlsx2)

        wb1 = load_workbook(xlsx1, read_only=True, data_only=True)
        wb2 = load_workbook(xlsx2, read_only=True, data_only=True)
        vals1 = list(wb1["audio"].iter_rows(min_row=1, max_row=1, values_only=True))[0]
        vals2 = list(wb2["audio"].iter_rows(min_row=1, max_row=1, values_only=True))[0]
        wb1.close()
        wb2.close()

        assert vals1 == vals2


# ── Multi-sheet: long audio (> 16 384 samples) ───────────────────────────


class TestMultiSheet:
    """Audio longer than EXCEL_MAX_COLS must span multiple sheets."""

    def test_wav_to_excel_creates_multiple_sheets(self, tmp_path):
        """wav_to_excel splits long audio across audio, audio_2, etc."""
        n_samples = EXCEL_MAX_COLS + 5000
        audio = _sine(n_samples / 16000, sr=16000)
        wav_in = _make_wav(tmp_path, audio)
        xlsx = tmp_path / "long.xlsx"

        result = wav_to_excel(wav_in, xlsx)
        assert result.num_samples == n_samples

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        assert "audio" in wb.sheetnames
        assert "audio_2" in wb.sheetnames

        row1 = list(wb["audio"].iter_rows(min_row=1, max_row=1, values_only=True))[0]
        row2 = list(wb["audio_2"].iter_rows(min_row=1, max_row=1, values_only=True))[0]
        non_none_1 = [v for v in row1 if v is not None]
        non_none_2 = [v for v in row2 if v is not None]
        assert len(non_none_1) == EXCEL_MAX_COLS
        assert len(non_none_2) == 5000
        wb.close()

    def test_round_trip_long_audio(self, tmp_path):
        """Round-trip with >16 384 samples preserves all data."""
        n_samples = EXCEL_MAX_COLS * 2 + 100
        audio = _sine(n_samples / 16000, sr=16000)
        wav_in = _make_wav(tmp_path, audio)
        xlsx = tmp_path / "long.xlsx"
        wav_out = tmp_path / "long_rt.wav"

        wav_to_excel(wav_in, xlsx)
        result = excel_to_wav(xlsx, wav_out, sample_rate=16000)

        assert result.num_samples == n_samples

        recovered, _ = load_audio(wav_out)
        assert len(recovered) == n_samples
        np.testing.assert_allclose(recovered, audio, atol=1e-4)

    def test_three_sheets(self, tmp_path):
        """Audio spanning 3 sheets is correctly named."""
        n_samples = EXCEL_MAX_COLS * 3
        audio = _sine(n_samples / 16000, sr=16000)
        wav_in = _make_wav(tmp_path, audio)
        xlsx = tmp_path / "three.xlsx"

        wav_to_excel(wav_in, xlsx)

        wb = load_workbook(xlsx, read_only=True, data_only=True)
        assert wb.sheetnames[:3] == ["audio", "audio_2", "audio_3"]
        wb.close()


# ── Pipeline Excel ────────────────────────────────────────────────────────


class TestPipelineExcel:
    """Tests for wav_to_pipeline_excel (requires soundperception)."""

    def test_pipeline_audio_only(self, tmp_path):
        """Pipeline with only audio export works and is readable by excel_to_wav."""
        audio = _sine(0.5)
        wav_in = _make_wav(tmp_path, audio)
        xlsx = tmp_path / "pipeline.xlsx"
        wav_out = tmp_path / "pipeline_rt.wav"

        result = wav_to_pipeline_excel(
            wav_in, xlsx,
            export_audio=True,
            export_periphery=False,
            export_integration=False,
        )
        assert result.num_samples == len(audio)

        result2 = excel_to_wav(xlsx, wav_out, sample_rate=16000)
        assert result2.num_samples == len(audio)

        recovered, _ = load_audio(wav_out)
        np.testing.assert_allclose(recovered, audio, atol=1e-4)

    def test_pipeline_with_periphery(self, tmp_path):
        """Pipeline with periphery creates memo_ma and memo_ck sheets."""
        audio = _sine(0.1)
        wav_in = _make_wav(tmp_path, audio)
        xlsx = tmp_path / "pipeline_periph.xlsx"

        wav_to_pipeline_excel(
            wav_in, xlsx,
            export_audio=True,
            export_periphery=True,
            export_integration=False,
        )

        wb = load_workbook(xlsx, read_only=True)
        assert "audio" in wb.sheetnames
        assert "memo_ma" in wb.sheetnames
        assert "memo_ck" in wb.sheetnames
        wb.close()

    def test_pipeline_excel_to_wav_reads_audio_not_periphery(self, tmp_path):
        """excel_to_wav on a pipeline file reads 'audio' sheet, not memo_ma."""
        audio = _sine(0.1)
        wav_in = _make_wav(tmp_path, audio)
        xlsx = tmp_path / "pipeline_multi.xlsx"
        wav_out = tmp_path / "from_pipeline.wav"

        wav_to_pipeline_excel(
            wav_in, xlsx,
            export_audio=True,
            export_periphery=True,
        )

        result = excel_to_wav(xlsx, wav_out, sample_rate=16000)
        assert result.num_samples == len(audio)

        recovered, _ = load_audio(wav_out)
        np.testing.assert_allclose(recovered, audio, atol=1e-4)

    def test_pipeline_long_audio_multi_sheet(self, tmp_path):
        """Pipeline with long audio creates audio + audio_2 sheets."""
        n_samples = EXCEL_MAX_COLS + 1000
        audio = _sine(n_samples / 16000, sr=16000)
        wav_in = _make_wav(tmp_path, audio)
        xlsx = tmp_path / "pipeline_long.xlsx"
        wav_out = tmp_path / "pipeline_long_rt.wav"

        wav_to_pipeline_excel(
            wav_in, xlsx,
            export_audio=True,
            export_periphery=False,
        )

        wb = load_workbook(xlsx, read_only=True)
        assert "audio" in wb.sheetnames
        assert "audio_2" in wb.sheetnames
        wb.close()

        result = excel_to_wav(xlsx, wav_out, sample_rate=16000)
        assert result.num_samples == n_samples


# ── Edge cases ────────────────────────────────────────────────────────────


class TestEdgeCases:
    def test_empty_excel_raises(self, tmp_path):
        """excel_to_wav raises ConversionError on empty Excel."""
        from openpyxl import Workbook
        xlsx = tmp_path / "empty.xlsx"
        wb = Workbook()
        wb.save(xlsx)

        with pytest.raises(ConversionError, match="Aucune donnée"):
            excel_to_wav(xlsx, tmp_path / "out.wav", sample_rate=16000)

    def test_no_export_selected_raises(self, tmp_path):
        """Pipeline with all exports disabled raises ConversionError."""
        audio = _sine(0.1)
        wav_in = _make_wav(tmp_path, audio)

        with pytest.raises(ConversionError, match="Aucune étape"):
            wav_to_pipeline_excel(
                wav_in, tmp_path / "nope.xlsx",
                export_audio=False,
                export_periphery=False,
                export_integration=False,
            )

    def test_peak_positive_sample(self, tmp_path):
        """A sample at +1.0 should not wrap to negative after int16 cast."""
        samples = np.array([1.0, -1.0, 0.0], dtype=np.float32)
        wav_in = _make_wav(tmp_path, samples)
        xlsx = tmp_path / "peak.xlsx"
        wav_out = tmp_path / "peak_rt.wav"

        wav_to_excel(wav_in, xlsx)
        excel_to_wav(xlsx, wav_out, sample_rate=16000)

        recovered, _ = load_audio(wav_out)
        assert recovered[0] > 0, f"Peak +1.0 became {recovered[0]} (sign flip!)"
        assert recovered[1] < 0, f"Peak -1.0 became {recovered[1]} (sign flip!)"
